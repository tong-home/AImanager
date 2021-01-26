from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# 在A1内输入42
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("C:\\Users\\admin\\Desktop\\sample.xlsx")
