#coding:utf-8
#************************
#author：wmm
#date：2019-7-10
#需要人工把解析为汉字的房间号手动修改为数字
#如果有识别为空的内容，手动添加为无字，否则报错
#*************************
#原理：编辑距离，又称Levenshtein距离（莱文斯坦距离也叫做Edit Distance），是指两个字串之间，由一个转成另一个所需的最少编辑操作次数，如果它们的距离越大，说明它们越是不同。许可的编辑操作包括将一个字符替换成另一个字符，插入一个字符，删除一个字符

from openpyxl import load_workbook
import openpyxl

def minEditDist(sm,sn):
    m = len(sm) + 1
    n  = len(sn) + 1

    # create a matrix (m*n)
    matrix = [[0] * n for i in range(m)]

    matrix[0][0] = 0
    for i in range(1, m):
        matrix[i][0] = matrix[i - 1][0] + 1

    for j in range(1, n):
        matrix[0][j] = matrix[0][j - 1] + 1

    for i in range(m):
        print(matrix[i])

    print("********************")

    cost = 0

    for i in range(1, m):
        for j in range(1, n):
            if sm[i - 1] == sn[j - 1]:
                cost = 0
            else:
                cost = 1

            matrix[i][j] = min(matrix[i - 1][j] + 1, matrix[i][j - 1] + 1, matrix[i - 1][j - 1] + cost)

    for i in range(m):
        print(matrix[i])


    return matrix[m - 1][n - 1]



#打开文件，最好使用绝对路径\\,否则可能报错
wb = load_workbook('C:\\Users\\admin\\Desktop\\潘多拉生产环境测试\\AAA语音识别测试\\众荟测试-asr&nlu_终板.xlsx')
sheetname=wb.get_sheet_by_name('测试用例与识别结果比对')
i=2
#循环取出测试用例中的内容
while(i<537):
    #取出表格第i行，第6列的单元格
    a=sheetname.cell(row=i, column=6)
    b=sheetname.cell(row=i, column=7)
    # 取出表格第i行，第6列的单元格中的值
    a=a.value
    b=b.value
    c=sheetname.cell(row=i,column=8)
    mindist = minEditDist(a, b)
    c.value=mindist
    print(c.value)
    #将文件保存到桌面的某个位置C:\Users\admin\Desktop\tt.xlsx
    wb.save(r'C:\Users\admin\Desktop\tt.xlsx')
    i = i + 1
#wb.save(r'C:\Users\admin\Desktop\tt.xlsx')
print('%s条测试用例全部计算结束'%(i-2))
